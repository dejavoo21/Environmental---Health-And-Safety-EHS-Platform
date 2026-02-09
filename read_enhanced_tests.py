import openpyxl

file = r"EHS_Enhanced_System_Tests\EHS_Enhanced_System_Tests.xlsx"
wb = openpyxl.load_workbook(file)

print("Sheet names:", wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 20:
            print(row)
        else:
            print(f"... ({ws.max_row - 19} more rows)")
            break
