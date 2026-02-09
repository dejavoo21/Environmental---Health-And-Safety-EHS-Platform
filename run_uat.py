import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import json

BASE_URL = "http://127.0.0.1:3001/api"
CREDENTIALS = {
    "admin": {"email": "admin@ehs.local", "password": "Admin123!"},
    "manager": {"email": "manager@ehs.local", "password": "Manager123!"},
    "worker": {"email": "worker@ehs.local", "password": "Worker123!"}
}

PASS_COLOR = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
FAIL_COLOR = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")
PASS_FONT = Font(bold=True, color="006400")
FAIL_FONT = Font(bold=True, color="8B0000")

tokens = {}

def login(role):
    try:
        creds = CREDENTIALS.get(role)
        r = requests.post(f"{BASE_URL}/auth/login", json=creds, timeout=5)
        if r.status_code == 200:
            token = r.json().get('data', {}).get('token')
            if token:
                tokens[role] = token
                return True, "OK"
        return False, f"Status: {r.status_code}"
    except Exception as e:
        return False, str(e)

# Test P1-01: Login
print("Testing P1-01: Login...")
results = {}
for role in ["admin", "manager", "worker"]:
    success, msg = login(role)
    results[role] = success
p1_01 = ("Pass" if all(results.values()) else "Fail", f"Admin:{results['admin']}, Manager:{results['manager']}, Worker:{results['worker']}")

# Test P1-02: Create Incident
print("Testing P1-02: Create Incident...")
try:
    if "worker" not in tokens:
        login("worker")
    headers = {"Authorization": f"Bearer {tokens['worker']}"}
    data = {"title": f"UAT-{datetime.now().timestamp()}", "description": "Test", "severity": "medium", "incident_type_id": 1, "site_id": None}
    r = requests.post(f"{BASE_URL}/incidents", json=data, headers=headers, timeout=5)
    incident_created = r.status_code in [200, 201]
    p1_02 = ("Pass" if incident_created else "Fail", f"Status: {r.status_code}")
except Exception as e:
    p1_02 = ("Fail", str(e))

# Test P1-03: Update Status
print("Testing P1-03: Update Status...")
try:
    if "manager" not in tokens:
        login("manager")
    headers = {"Authorization": f"Bearer {tokens['manager']}"}
    r = requests.get(f"{BASE_URL}/incidents?limit=1", headers=headers, timeout=5)
    if r.status_code == 200 and r.json().get('data'):
        incident_id = r.json()['data'][0]['id']
        update_r = requests.put(f"{BASE_URL}/incidents/{incident_id}", json={"status": "closed"}, headers=headers, timeout=5)
        status_updated = update_r.status_code in [200, 204]
        p1_03 = ("Pass" if status_updated else "Fail", f"Update status: {update_r.status_code}")
    else:
        p1_03 = ("Fail", "No incidents")
except Exception as e:
    p1_03 = ("Fail", str(e))

# Test P1-04: Dashboard
print("Testing P1-04: Dashboard...")
try:
    if "manager" not in tokens:
        login("manager")
    headers = {"Authorization": f"Bearer {tokens['manager']}"}
    r = requests.get(f"{BASE_URL}/dashboard/summary", headers=headers, timeout=5)
    has_data = r.status_code == 200 and 'data' in r.json()
    p1_04 = ("Pass" if has_data else "Fail", f"Status: {r.status_code}")
except Exception as e:
    p1_04 = ("Fail", str(e))

print("\nResults:")
print(f"P1-01: {p1_01[0]} - {p1_01[1]}")
print(f"P1-02: {p1_02[0]} - {p1_02[1]}")
print(f"P1-03: {p1_03[0]} - {p1_03[1]}")
print(f"P1-04: {p1_04[0]} - {p1_04[1]}")

# Update Excel files
results_map = {
    "P1-01": p1_01,
    "P1-02": p1_02,
    "P1-03": p1_03,
    "P1-04": p1_04,
}

print("\nUpdating Excel files...")

# Update Signoff_v1
signoff_file = r"EHS_UAT_Signoff_v1\EHS_UAT_Signoff_v1.xlsx"
wb = openpyxl.load_workbook(signoff_file)
ws = wb["UAT Sign-Off"]

updated_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    test_id = row[1].value
    if test_id in results_map:
        result, notes = results_map[test_id]
        row[5].value = result
        row[5].fill = PASS_COLOR if result == "Pass" else FAIL_COLOR
        row[5].font = PASS_FONT if result == "Pass" else FAIL_FONT
        row[6].value = notes
        updated_count += 1

wb.save(signoff_file)
print(f"Updated Signoff_v1: {updated_count} tests")

# Update Phase5
phase5_file = r"EHS_UAT_Phase5_Analytics\EHS_UAT_Phase5_Analytics.xlsx"
wb = openpyxl.load_workbook(phase5_file)
ws = wb["Phase 5 UAT"]

updated_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    test_id = row[1].value
    if test_id in results_map:
        result, notes = results_map[test_id]
        row[5].value = result
        row[5].fill = PASS_COLOR if result == "Pass" else FAIL_COLOR
        row[5].font = PASS_FONT if result == "Pass" else FAIL_FONT
        row[6].value = notes
        updated_count += 1

wb.save(phase5_file)
print(f"Updated Phase5: {updated_count} tests")

print("\nDone!")
