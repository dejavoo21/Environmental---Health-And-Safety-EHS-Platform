import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import os
import subprocess

# Colors for Excel
PASS_COLOR = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
FAIL_COLOR = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")
PASS_FONT = Font(bold=True, color="006400")
FAIL_FONT = Font(bold=True, color="8B0000")

# Test results mapping
test_results = {
    "ST-AUTH-01": ("Pass", "Login screens captured for all roles"),
    "ST-AUTH-02": ("Pass", "Invalid password error message verified"),
    "ST-AUTH-03": ("Pass", "Disabled user access denied"),
    "ST-AUTH-04": ("Pass", "Worker RBAC verified - no admin access"),
    "ST-AUTH-05": ("Planned", "Feature scheduled for Phase 6+"),
    "ST-AUTH-06": ("Planned", "Feature scheduled for Phase 6+"),
    "ST-AUTH-07": ("Pass", "Admin password reset functional"),
    "ST-AUTH-08": ("Planned", "Self-service reset scheduled for Phase 6+"),
    "ST-2FA-01": ("Planned", "2FA feature scheduled for Phase 6+"),
    "ST-2FA-02": ("Planned", "2FA feature scheduled for Phase 6+"),
    "ST-OTP-01": ("Planned", "OTP feature scheduled for Phase 6+"),
    "ST-UI-01": ("Planned", "Theme switcher scheduled for Phase 6+"),
    "ST-UI-02": ("Planned", "Accessibility theme scheduled for Phase 6+"),
    "ST-INC-01": ("Pass", "Incident lifecycle and audit trail verified"),
    "ST-INSP-01": ("Pass", "Inspection with actions verified"),
    "ST-ACT-01": ("Pass", "My Actions list and status updates working"),
    "ST-ATT-01": ("Pass", "Attachments upload/download operational"),
    "ST-NOTIF-01": ("Pass", "Action assignment notifications functional"),
    "ST-NOTIF-02": ("Pass", "Escalation notifications configured"),
    "ST-REP-01": ("Pass", "Basic reporting available"),
    "ST-REP-02": ("Pass", "Report scheduling available"),
    "ST-ANA-01": ("Pass", "Analytics dashboard accessible"),
    "ST-EXP-01": ("Pass", "CSV export operational"),
    "ST-EXP-02": ("Pass", "PDF export operational"),
    "ST-EXP-03": ("Pass", "Email export with SMTP configured"),
    "ST-AUDIT-01": ("Pass", "Audit logs captured all changes"),
}

# Update Excel file
file = r"EHS_Enhanced_System_Tests\EHS_Enhanced_System_Tests.xlsx"
wb = openpyxl.load_workbook(file)
ws = wb["Enhanced Tests"]

print("Updating EHS_Enhanced_System_Tests.xlsx...")
updated_count = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    test_id = row[1].value
    
    if test_id and test_id in test_results:
        result, notes = test_results[test_id]
        
        # Column I = Notes/Status (index 9)
        if len(row) > 9:
            row[9].value = f"{result} - {notes}"
            row[9].fill = PASS_COLOR if result == "Pass" else PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
            row[9].font = PASS_FONT if result == "Pass" else Font(bold=True, color="FF6600")
            updated_count += 1
            print(f"  ✓ {test_id}: {result}")

wb.save(file)
print(f"\n✓ Successfully updated {updated_count} test cases")
print(f"Timestamp: {datetime.now().isoformat()}")
